import matplotlib.pyplot as plt
import xlrd as rd
import pandas as pd
import numpy as np
import openpyxl as op
import os
from ReadFile import readFile
from findPoints import Find_Points
from weakStudents import weakStudents
from complexGraphs import complexGraphs

class CheckData:

    def __init__(self,sec):
        self.sec=sec

    def check_SecifExists(self,file):
        if (os.path.isfile(file+'.xlsx')):
            filename = file + '.xlsx'
            wb = op.load_workbook(filename)
            sheets=wb.sheetnames
            for i in range (len(sheets)):
                if(self.sec==sheets[i]):
                    return True
            return False
        else:
            return False

    def check_colifExist(self,col,file):
        if (os.path.isfile(file + '.xlsx')):
             filename=file+'.xlsx'
             wb = op.load_workbook(filename)
             sheet=wb[self.sec]
             for rowOfCellObjects in sheet['I1':'CO1']:
                    for cellObj in rowOfCellObjects:
                        #print(cellObj.coordinate, cellObj.value)
                        if(col==cellObj.value):
                                return True
             return False
        else:
            return False
    def check_ifDataExist(self,col,file):
        if (os.path.isfile(file + '.xlsx')):
            filename = file + '.xlsx'
            sheet = pd.read_excel(filename, self.sec)
            colum=sheet[col]
            length=len(sheet[col])
            check=False
            for i in range(length):
                if(colum[i] is not np.nan):
                    check = True

            return check
        else:
            return False
